"""Report generation service — dynamic PDF and Excel exports."""

import io
import json
from datetime import datetime, timezone
from typing import Optional, Dict, Any, List

from fastapi import HTTPException, status
from sqlalchemy.orm import Session
from sqlalchemy import text


class ReportService:
    """Generates downloadable PDF and Excel reports from persistent prediction records."""

    @staticmethod
    def _get_prediction_dict(db: Session, prediction_id: Optional[str] = None, project_id: Optional[str] = None) -> Dict[str, Any]:
        """Fetch prediction record directly from repository_predictions and repositories tables."""
        row = None
        if prediction_id:
            query = text("""
                SELECT p.id as prediction_id, p.prediction_status, p.model_version, p.triggered_by, p.created_at,
                       p.repository_id, p.failure_probability, p.risk_score, p.risk_level, p.confidence, p.health_score,
                       p.feature_importance_json, p.recommendations_json,
                       r.repository_name, r.repository_url, r.organization, r.language, r.git_provider, r.branch, r.visibility, r.owner
                FROM repository_predictions p
                LEFT JOIN repositories r ON p.repository_id = r.id
                WHERE p.id = :pid
            """)
            row = db.execute(query, {"pid": prediction_id}).mappings().first()

        if not row and project_id:
            query = text("""
                SELECT p.id as prediction_id, p.prediction_status, p.model_version, p.triggered_by, p.created_at,
                       p.repository_id, p.failure_probability, p.risk_score, p.risk_level, p.confidence, p.health_score,
                       p.feature_importance_json, p.recommendations_json,
                       r.repository_name, r.repository_url, r.organization, r.language, r.git_provider, r.branch, r.visibility, r.owner
                FROM repository_predictions p
                LEFT JOIN repositories r ON p.repository_id = r.id
                WHERE p.repository_id = :rid
                ORDER BY p.created_at DESC
            """)
            row = db.execute(query, {"rid": project_id}).mappings().first()

        if not row:
            # Fallback query if no record match
            query = text("""
                SELECT p.id as prediction_id, p.prediction_status, p.model_version, p.triggered_by, p.created_at,
                       p.repository_id, p.failure_probability, p.risk_score, p.risk_level, p.confidence, p.health_score,
                       p.feature_importance_json, p.recommendations_json,
                       r.repository_name, r.repository_url, r.organization, r.language, r.git_provider, r.branch, r.visibility, r.owner
                FROM repository_predictions p
                LEFT JOIN repositories r ON p.repository_id = r.id
                ORDER BY p.created_at DESC
            """)
            row = db.execute(query).mappings().first()

        if not row:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Prediction data not found for ID: {prediction_id or project_id}")

        res = dict(row)
        
        # Load additional metrics if repository_id is present
        repo_id = res.get("repository_id")
        if repo_id:
            metrics_query = text("""
                SELECT commit_count, commit_frequency, pull_requests, merged_pull_requests, failed_pull_requests,
                       contributors, active_contributors, inactive_days, open_issues, closed_issues,
                       code_coverage, documentation_score, build_success_rate, cyclomatic_complexity,
                       technical_debt, bus_factor, velocity, updated_at
                FROM repository_metrics
                WHERE repository_id = :rid
            """)
            metrics_row = db.execute(metrics_query, {"rid": repo_id}).mappings().first()
            if metrics_row:
                res["metrics"] = dict(metrics_row)
            else:
                res["metrics"] = {}
        else:
            res["metrics"] = {}

        return res

    @staticmethod
    def generate_pdf(
        db: Session,
        user_id: str,
        prediction_id: Optional[str] = None,
        project_id: Optional[str] = None,
        ip_address: Optional[str] = None,
    ) -> bytes:
        data = ReportService._get_prediction_dict(db, prediction_id, project_id)

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, HRFlowable
        except ImportError:
            raise HTTPException(status_code=500, detail="reportlab library is required for PDF generation.")

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle(
            'ReportTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, leading=24, textColor=colors.HexColor('#0F172A')
        )
        subtitle_style = ParagraphStyle(
            'ReportSub', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, textColor=colors.HexColor('#64748B')
        )
        section_style = ParagraphStyle(
            'ReportSec', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=13, leading=16, textColor=colors.HexColor('#1E293B'), spaceBefore=12, spaceAfter=6
        )

        elements.append(Paragraph("RIVEXA — Predictive Risk Intelligence Platform", title_style))
        elements.append(Paragraph(f"Repository Risk Prediction Report — Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", subtitle_style))
        elements.append(Spacer(1, 10))
        elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#E2E8F0'), spaceAfter=15))

        # ── 1. Executive Prediction Summary ───────────────────────────────────
        elements.append(Paragraph("1. Executive Prediction Summary", section_style))

        fail_prob = data.get('failure_probability') if data.get('failure_probability') is not None else 0.0
        risk_score = data.get('risk_score') if data.get('risk_score') is not None else 0
        risk_level = data.get('risk_level') or 'LOW'
        confidence = data.get('confidence') if data.get('confidence') is not None else 0.0
        health_score = data.get('health_score') if data.get('health_score') is not None else 0.0

        summary_table_data = [
            ["Prediction ID", str(data.get('prediction_id') or 'N/A'), "Risk Level", f"{risk_level} RISK"],
            ["Failure Probability", f"{fail_prob * 100:.1f}%", "Risk Score", f"{risk_score} / 100"],
            ["AI Confidence", f"{confidence * 100:.1f}%", "Health Score", f"{health_score:.1f} / 100"],
            ["Model Version", str(data.get('model_version') or 'RandomForest-v1.0'), "Prediction Status", str(data.get('prediction_status') or 'COMPLETED')],
            ["Triggered By", str(data.get('triggered_by') or 'MANUAL'), "Predicted At", str(data.get('created_at') or 'N/A')]
        ]

        t_summary = Table(summary_table_data, colWidths=[130, 140, 130, 140])
        t_summary.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F8FAFC')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F8FAFC')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1E293B')),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        elements.append(t_summary)
        elements.append(Spacer(1, 15))

        # ── 2. Repository Metadata ───────────────────────────────────────────
        elements.append(Paragraph("2. Repository Metadata", section_style))
        repo_table_data = [
            ["Repository Name", str(data.get('repository_name') or 'N/A'), "Repository ID", str(data.get('repository_id') or 'N/A')],
            ["Git Provider", str(data.get('git_provider') or 'GITHUB'), "Branch", str(data.get('branch') or 'main')],
            ["Visibility", str(data.get('visibility') or 'PUBLIC'), "Language", str(data.get('language') or 'N/A')],
            ["Organization / Owner", str(data.get('organization') or data.get('owner') or 'N/A'), "Repository URL", str(data.get('repository_url') or 'N/A')]
        ]
        t_repo = Table(repo_table_data, colWidths=[130, 140, 130, 140])
        t_repo.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#F1F5F9')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#F1F5F9')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
        ]))
        elements.append(t_repo)
        elements.append(Spacer(1, 15))

        # ── 3. Repository Performance Metrics ──────────────────────────────────
        metrics = data.get("metrics") or {}
        if metrics:
            elements.append(Paragraph("3. Repository Performance Metrics", section_style))
            metrics_table_data = [
                ["Metric Name", "Value", "Metric Name", "Value"],
                ["Total Commits", str(metrics.get('commit_count') or 0), "Commit Frequency", f"{metrics.get('commit_frequency', 0.0):.2f}/wk"],
                ["Contributors", str(metrics.get('contributors') or 0), "Active Contributors", str(metrics.get('active_contributors') or 0)],
                ["Open Issues", str(metrics.get('open_issues') or 0), "Closed Issues", str(metrics.get('closed_issues') or 0)],
                ["Pull Requests", str(metrics.get('pull_requests') or 0), "Merged PRs", str(metrics.get('merged_pull_requests') or 0)],
                ["Code Coverage", f"{metrics.get('code_coverage', 0.0):.1f}%", "Technical Debt Ratio", f"{metrics.get('technical_debt', 0.0):.1f}%"],
                ["Build Success Rate", f"{metrics.get('build_success_rate', 0.0):.1f}%", "Bus Factor Strength", str(metrics.get('bus_factor') or 1)],
                ["Inactive Period", f"{metrics.get('inactive_days') or 0} days", "Code Churn Velocity", f"{metrics.get('velocity', 0.0):.1f}"]
            ]
            t_metrics = Table(metrics_table_data, colWidths=[130, 140, 130, 140])
            t_metrics.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F8FAFC')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            elements.append(t_metrics)
            elements.append(Spacer(1, 15))

        # ── 4. SHAP Feature Importance & Explainability ──────────────────────
        elements.append(Paragraph("4. SHAP Feature Importance & Explainability", section_style))
        shap_raw = data.get('feature_importance_json')
        shap_list = []
        if shap_raw:
            try:
                shap_list = json.loads(shap_raw) if isinstance(shap_raw, str) else shap_raw
            except Exception:
                shap_list = []

        if shap_list and isinstance(shap_list, list):
            shap_table_data = [["Feature Name", "Impact Value", "Risk Direction"]]
            for item in shap_list[:10]:
                feat_name = item.get('display_name') or item.get('feature_name') or item.get('feature') or 'Feature'
                imp_val = f"{float(item.get('impact', 0)):.4f}"
                direction = str(item.get('direction', 'increases_risk')).replace('_', ' ').title()
                shap_table_data.append([feat_name, imp_val, direction])

            t_shap = Table(shap_table_data, colWidths=[200, 140, 200])
            t_shap.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F172A')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            elements.append(t_shap)
        else:
            elements.append(Paragraph("No SHAP feature importance details available.", subtitle_style))

        elements.append(Spacer(1, 15))

        # ── 5. AI Recommendations ────────────────────────────────────────────
        elements.append(Paragraph("5. Strategic AI Recommendations", section_style))
        recs_raw = data.get('recommendations_json')
        recs_list = []
        if recs_raw:
            try:
                recs_list = json.loads(recs_raw) if isinstance(recs_raw, str) else recs_raw
            except Exception:
                recs_list = []

        is_rich = False
        recs_parsed = []
        if isinstance(recs_list, dict) and "recommendations" in recs_list:
            recs_parsed = recs_list["recommendations"]
            is_rich = True
        elif isinstance(recs_list, list):
            if recs_list and isinstance(recs_list[0], dict):
                recs_parsed = recs_list
                is_rich = True
            else:
                recs_parsed = recs_list

        if recs_parsed:
            if is_rich:
                for idx, rec in enumerate(recs_parsed, start=1):
                    title = rec.get("title", "Recommendation")
                    action = rec.get("recommended_action", "")
                    why = rec.get("why_it_matters", "")
                    priority = rec.get("suggested_priority", "Medium")
                    
                    elements.append(Paragraph(f"<b>{idx}. {title}</b> (Priority: {priority})", ParagraphStyle('RecTitle', parent=styles['Normal'], fontSize=10, leading=14, fontName='Helvetica-Bold')))
                    if action:
                        elements.append(Paragraph(f"• <b>Action:</b> {action}", ParagraphStyle('RecAction', parent=styles['Normal'], fontSize=9, leading=13, spaceAfter=2)))
                    if why:
                        elements.append(Paragraph(f"• <b>Rationale:</b> {why}", ParagraphStyle('RecWhy', parent=styles['Normal'], fontSize=9, leading=13, spaceAfter=6)))
            else:
                for idx, rec in enumerate(recs_parsed, start=1):
                    elements.append(Paragraph(f"• <b>Recommendation {idx}:</b> {rec}", ParagraphStyle('Rec', parent=styles['Normal'], fontSize=9, leading=13, spaceAfter=4)))
        else:
            elements.append(Paragraph("Maintain active repository monitoring and continuous integration checks.", subtitle_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def generate_excel(
        db: Session,
        user_id: str,
        prediction_id: Optional[str] = None,
        project_id: Optional[str] = None,
        ip_address: Optional[str] = None,
    ) -> bytes:
        data = ReportService._get_prediction_dict(db, prediction_id, project_id)

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl library is required for Excel generation.")

        wb = openpyxl.Workbook()

        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='0F172A', end_color='0F172A', fill_type='solid')
        title_font = Font(name='Calibri', bold=True, size=14, color='0F172A')

        # ── SHEET 1: Risk Summary ─────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Risk Summary"

        ws1["A1"] = "RIVEXA — Project Risk Summary"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"

        headers1 = ["Field", "Value"]
        for col_num, h in enumerate(headers1, 1):
            cell = ws1.cell(row=4, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill

        fail_prob = data.get('failure_probability') if data.get('failure_probability') is not None else 0.0
        confidence = data.get('confidence') if data.get('confidence') is not None else 0.0

        summary_rows = [
            ("Repository", str(data.get('repository_name') or 'N/A')),
            ("Owner", str(data.get('organization') or data.get('owner') or 'N/A')),
            ("URL", str(data.get('repository_url') or 'N/A')),
            ("Language", str(data.get('language') or 'N/A')),
            ("Prediction Date", str(data.get('created_at') or 'N/A')),
            ("Risk Level", str(data.get('risk_level') or 'LOW')),
            ("Risk Score", data.get('risk_score') if data.get('risk_score') is not None else int(round(fail_prob * 100))),
            ("Confidence", f"{confidence * 100:.1f}%"),
            ("Project Status", str(data.get('prediction_status') or 'COMPLETED'))
        ]

        for idx, (field, value) in enumerate(summary_rows, start=5):
            ws1.cell(row=idx, column=1, value=field)
            ws1.cell(row=idx, column=2, value=value)

        ws1.column_dimensions["A"].width = 25
        ws1.column_dimensions["B"].width = 50

        # ── SHEET 2: Risk Analysis ────────────────────────────────────────────
        ws2 = wb.create_sheet(title="Risk Analysis")
        ws2["A1"] = "Risk Factor Analysis"
        ws2["A1"].font = title_font

        headers2 = ["Risk category", "Severity", "Detected issue", "Evidence", "Impact"]
        for col_num, h in enumerate(headers2, 1):
            cell = ws2.cell(row=3, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill

        shap_raw = data.get('feature_importance_json')
        shap_list = []
        if shap_raw:
            try:
                shap_list = json.loads(shap_raw) if isinstance(shap_raw, str) else shap_raw
            except Exception:
                shap_list = []

        if shap_list and isinstance(shap_list, list):
            for idx, item in enumerate(shap_list, start=4):
                feat_name = item.get('feature_name') or item.get('feature') or 'feature'
                display_name = item.get('display_name') or feat_name
                imp_val = item.get('impact', 0.0)
                
                # Determine severity
                abs_imp = abs(imp_val)
                if abs_imp > 0.25:
                    severity = "High"
                elif abs_imp > 0.1:
                    severity = "Medium"
                else:
                    severity = "Low"

                # Determine risk category
                feat_lower = feat_name.lower()
                if "bus" in feat_lower:
                    category = "Team Distribution / Bus Factor"
                elif "commit" in feat_lower or "velocity" in feat_lower or "inactive" in feat_lower:
                    category = "Repository Activity"
                elif "contributor" in feat_lower:
                    category = "Community Health"
                elif "issue" in feat_lower:
                    category = "Issue Management"
                elif "coverage" in feat_lower or "documentation" in feat_lower:
                    category = "Code Quality"
                elif "build" in feat_lower:
                    category = "CI/CD Build Health"
                elif "debt" in feat_lower or "complexity" in feat_lower:
                    category = "Technical Debt"
                else:
                    category = "General Code Metrics"

                # Determine impact direction and statement
                if imp_val > 0:
                    impact = "Increases project failure risk probability"
                    evidence = f"+{imp_val:.4f} SHAP impact"
                else:
                    impact = "Mitigates project risk and improves stability"
                    evidence = f"{imp_val:.4f} SHAP impact"

                ws2.cell(row=idx, column=1, value=category)
                ws2.cell(row=idx, column=2, value=severity)
                ws2.cell(row=idx, column=3, value=display_name)
                ws2.cell(row=idx, column=4, value=evidence)
                ws2.cell(row=idx, column=5, value=impact)
        else:
            ws2.cell(row=4, column=1, value="No risk analysis details available")

        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 15
        ws2.column_dimensions["C"].width = 30
        ws2.column_dimensions["D"].width = 20
        ws2.column_dimensions["E"].width = 45

        # ── SHEET 3: Repository Metrics ───────────────────────────────────────
        ws3 = wb.create_sheet(title="Repository Metrics")
        ws3["A1"] = "Extracted GitHub Repository Metrics"
        ws3["A1"].font = title_font

        headers3 = ["Metric", "Value"]
        for col_num, h in enumerate(headers3, 1):
            cell = ws3.cell(row=3, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill

        metrics = data.get("metrics") or {}
        
        metrics_rows = [
            ("Stars", 0),
            ("Forks", 0),
            ("Open Issues", metrics.get('open_issues', 0)),
            ("Closed Issues", metrics.get('closed_issues', 0)),
            ("Total Commits", metrics.get('commit_count', 0)),
            ("Commit Frequency", f"{metrics.get('commit_frequency', 0.0):.2f} commits/week"),
            ("Contributors", metrics.get('contributors', 0)),
            ("Active Contributors", metrics.get('active_contributors', 0)),
            ("Inactive Period (Days)", metrics.get('inactive_days', 0)),
            ("Pull Requests", metrics.get('pull_requests', 0)),
            ("Merged Pull Requests", metrics.get('merged_pull_requests', 0)),
            ("Failed Pull Requests", metrics.get('failed_pull_requests', 0)),
            ("Code Coverage", f"{metrics.get('code_coverage', 0.0):.1f}%"),
            ("Documentation Score", f"{metrics.get('documentation_score', 0.0):.1f}%"),
            ("Build Success Rate", f"{metrics.get('build_success_rate', 0.0):.1f}%"),
            ("Cyclomatic Complexity", f"{metrics.get('cyclomatic_complexity', 0.0):.2f}"),
            ("Technical Debt Ratio", f"{metrics.get('technical_debt', 0.0):.1f}%"),
            ("Bus Factor", metrics.get('bus_factor', 1)),
            ("Code Churn Velocity", f"{metrics.get('velocity', 0.0):.2f}")
        ]

        for idx, (metric, val) in enumerate(metrics_rows, start=4):
            ws3.cell(row=idx, column=1, value=metric)
            ws3.cell(row=idx, column=2, value=val)

        ws3.column_dimensions["A"].width = 30
        ws3.column_dimensions["B"].width = 25

        # ── SHEET 4: AI Recommendations ───────────────────────────────────────
        ws4 = wb.create_sheet(title="AI Recommendations")
        ws4["A1"] = "Strategic Risk Mitigation Plan"
        ws4["A1"].font = title_font

        headers4 = ["Risk", "Recommendation", "Priority", "Expected improvement", "Suggested action"]
        for col_num, h in enumerate(headers4, 1):
            cell = ws4.cell(row=3, column=col_num, value=h)
            cell.font = header_font
            cell.fill = header_fill

        recs_raw = data.get('recommendations_json')
        recs_list = []
        if recs_raw:
            try:
                recs_list = json.loads(recs_raw) if isinstance(recs_raw, str) else recs_raw
            except Exception:
                recs_list = []

        is_rich = False
        recs_parsed = []
        if isinstance(recs_list, dict) and "recommendations" in recs_list:
            recs_parsed = recs_list["recommendations"]
            is_rich = True
        elif isinstance(recs_list, list):
            if recs_list and isinstance(recs_list[0], dict):
                recs_parsed = recs_list
                is_rich = True
            else:
                recs_parsed = recs_list

        if recs_parsed:
            for idx, rec in enumerate(recs_parsed, start=4):
                if is_rich:
                    risk = rec.get("risk_detected", "General Failure Risk")
                    recommendation = rec.get("title", "Improve general metrics")
                    priority = rec.get("suggested_priority", "Medium")
                    expected_improvement = rec.get("estimated_risk_reduction", "Reduces project failure probability")
                    suggested_action = rec.get("recommended_action", "")
                else:
                    risk = "Legacy Risk Attributed"
                    recommendation = rec
                    priority = "Medium"
                    expected_improvement = "Reduces failure risk"
                    suggested_action = "Implement continuous monitoring"

                ws4.cell(row=idx, column=1, value=risk)
                ws4.cell(row=idx, column=2, value=recommendation)
                ws4.cell(row=idx, column=3, value=priority)
                ws4.cell(row=idx, column=4, value=expected_improvement)
                ws4.cell(row=idx, column=5, value=suggested_action)
        else:
            ws4.cell(row=4, column=1, value="Continuous Monitoring")
            ws4.cell(row=4, column=2, value="Maintain current development velocity and testing checks")
            ws4.cell(row=4, column=3, value="Low")
            ws4.cell(row=4, column=4, value="Mitigates regression risk")
            ws4.cell(row=4, column=5, value="Run daily build suites")

        ws4.column_dimensions["A"].width = 30
        ws4.column_dimensions["B"].width = 35
        ws4.column_dimensions["C"].width = 15
        ws4.column_dimensions["D"].width = 30
        ws4.column_dimensions["E"].width = 45

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
